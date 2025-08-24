#!/usr/bin/env python3
"""
HVAC Data Processor and Excel Generator
Processes scraped business data and generates professional Excel reports
"""

import json
import logging
import pandas as pd
import re
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional, Tuple
from dataclasses import asdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image

from hvac_scraper_core import BusinessInfo

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class HVACDataProcessor:
    """Advanced data processing and Excel generation for HVAC businesses"""
    
    def __init__(self):
        self.businesses: List[BusinessInfo] = []
        self.processed_data: pd.DataFrame = None
        
    def load_businesses(self, businesses: List[BusinessInfo]):
        """Load business data for processing"""
        self.businesses = businesses
        logger.info(f"Loaded {len(businesses)} businesses for processing")
    
    def clean_and_validate_data(self) -> pd.DataFrame:
        """Clean and validate business data"""
        logger.info("Starting data cleaning and validation")
        
        # Convert to DataFrame
        data = [asdict(business) for business in self.businesses]
        df = pd.DataFrame(data)
        
        if df.empty:
            logger.warning("No data to process")
            return df
        
        # Clean business names
        df['name'] = df['name'].apply(self._clean_business_name)
        
        # Validate and format phone numbers
        df['phone'] = df['phone'].apply(self._validate_phone_number)
        
        # Clean addresses
        df['address'] = df['address'].apply(self._clean_address)
        
        # Validate review counts and ratings
        df['review_count'] = pd.to_numeric(df['review_count'], errors='coerce').fillna(0).astype(int)
        df['star_rating'] = pd.to_numeric(df['star_rating'], errors='coerce').fillna(0.0)
        
        # Clean owner names
        df['owner_name'] = df['owner_name'].apply(self._clean_owner_name)
        
        # Remove duplicates based on name and phone
        initial_count = len(df)
        df = df.drop_duplicates(subset=['name', 'phone'], keep='first')
        removed_count = initial_count - len(df)
        
        if removed_count > 0:
            logger.info(f"Removed {removed_count} duplicate entries")
        
        # Filter out invalid entries
        df = df[df['name'].str.len() > 0]  # Must have a name
        
        # Add calculated fields
        df['priority_score'] = df.apply(self._calculate_priority_score, axis=1)
        df['data_quality_score'] = df.apply(self._calculate_data_quality_score, axis=1)
        
        self.processed_data = df
        logger.info(f"Data cleaning complete. {len(df)} valid businesses remaining")
        
        return df
    
    def _clean_business_name(self, name: str) -> str:
        """Clean and standardize business names"""
        if pd.isna(name) or not name:
            return ""
        
        name = str(name).strip()
        
        # Remove common prefixes/suffixes that might be artifacts
        artifacts = [
            r'^".*"$',  # Quoted strings
            r'^\d+\.\s*',  # Numbered lists
            r'Google Maps.*',  # Google Maps artifacts
        ]
        
        for pattern in artifacts:
            if re.match(pattern, name, re.IGNORECASE):
                return ""
        
        # Standardize business entity suffixes
        name = re.sub(r'\bLLC\.?$', 'LLC', name, flags=re.IGNORECASE)
        name = re.sub(r'\bInc\.?$', 'Inc', name, flags=re.IGNORECASE)
        name = re.sub(r'\bCorp\.?$', 'Corp', name, flags=re.IGNORECASE)
        
        return name.strip()
    
    def _validate_phone_number(self, phone: str) -> str:
        """Validate and format phone numbers"""
        if pd.isna(phone) or not phone:
            return ""
        
        phone = str(phone).strip()
        
        # Extract digits only
        digits = re.sub(r'\D', '', phone)
        
        # Validate US phone number format
        if len(digits) == 10:
            return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
        elif len(digits) == 11 and digits[0] == '1':
            return f"({digits[1:4]}) {digits[4:7]}-{digits[7:]}"
        
        return ""  # Invalid phone number
    
    def _clean_address(self, address: str) -> str:
        """Clean and standardize addresses"""
        if pd.isna(address) or not address:
            return ""
        
        address = str(address).strip()
        
        # Remove artifacts
        if address in ['Unknown', 'N/A', 'Not found', 'No reviews']:
            return ""
        
        # Basic address cleaning
        address = re.sub(r'\s+', ' ', address)  # Multiple spaces to single
        address = re.sub(r',\s*,', ',', address)  # Double commas
        
        return address.strip()
    
    def _clean_owner_name(self, owner: str) -> str:
        """Clean owner names"""
        if pd.isna(owner) or not owner:
            return ""
        
        owner = str(owner).strip()
        
        # Remove artifacts
        if owner in ['Unknown', 'N/A', 'Not found', 'Not explicitly stated']:
            return ""
        
        # Remove parenthetical information
        owner = re.sub(r'\s*\([^)]*\)', '', owner)
        
        # Remove titles
        owner = re.sub(r'^(Mr\.?|Mrs\.?|Ms\.?|Dr\.?)\s+', '', owner, flags=re.IGNORECASE)
        
        # Validate name format (should be 2-4 words, each starting with capital)
        words = owner.split()
        if len(words) < 2 or len(words) > 4:
            return ""
        
        # Check if all words look like names
        for word in words:
            if not re.match(r'^[A-Z][a-z]+$', word):
                return ""
        
        return owner.strip()
    
    def _calculate_priority_score(self, row: pd.Series) -> int:
        """Calculate priority score (lower = higher priority)"""
        score = 0
        
        # Review count (lower is better)
        score += row['review_count'] * 10
        
        # Has phone number (better)
        if row['phone']:
            score -= 50
        
        # Has owner name (better)
        if row['owner_name']:
            score -= 30
        
        # Has website (indicates established business)
        if row['website'] and row['website'] not in ['', 'Not found', 'N/A']:
            score -= 10
        
        # Star rating (higher is better for established businesses)
        if row['star_rating'] > 0:
            score -= int(row['star_rating'] * 2)
        
        return max(0, score)  # Ensure non-negative
    
    def _calculate_data_quality_score(self, row: pd.Series) -> float:
        """Calculate data quality score (0-100)"""
        score = 0
        max_score = 0
        
        # Required fields
        if row['name']:
            score += 20
        max_score += 20
        
        if row['phone']:
            score += 25
        max_score += 25
        
        if row['address']:
            score += 15
        max_score += 15
        
        # Optional but valuable fields
        if row['owner_name']:
            score += 20
        max_score += 20
        
        if row['website'] and row['website'] not in ['', 'Not found', 'N/A']:
            score += 10
        max_score += 10
        
        if row['review_count'] > 0:
            score += 5
        max_score += 5
        
        if row['additional_contact']:
            score += 5
        max_score += 5
        
        return (score / max_score) * 100 if max_score > 0 else 0
    
    def filter_by_criteria(self, max_reviews: int = 20, min_quality_score: float = 40.0) -> pd.DataFrame:
        """Filter businesses by specified criteria"""
        if self.processed_data is None:
            raise ValueError("No processed data available. Run clean_and_validate_data() first.")
        
        filtered = self.processed_data[
            (self.processed_data['review_count'] <= max_reviews) &
            (self.processed_data['data_quality_score'] >= min_quality_score) &
            (self.processed_data['phone'] != "")  # Must have phone number
        ].copy()
        
        # Sort by priority score (lowest first)
        filtered = filtered.sort_values(['priority_score', 'review_count', 'name'])
        
        logger.info(f"Filtered to {len(filtered)} businesses meeting criteria")
        return filtered
    
    def generate_excel_report(self, output_path: str, filtered_data: Optional[pd.DataFrame] = None) -> str:
        """Generate comprehensive Excel report"""
        logger.info(f"Generating Excel report: {output_path}")
        
        if filtered_data is None:
            filtered_data = self.filter_by_criteria()
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        self._create_summary_sheet(wb, filtered_data)
        self._create_calling_list_sheet(wb, filtered_data)
        self._create_zero_reviews_sheet(wb, filtered_data)
        self._create_location_breakdown_sheet(wb, filtered_data)
        self._create_data_quality_sheet(wb, filtered_data)
        
        # Save workbook
        wb.save(output_path)
        logger.info(f"Excel report saved: {output_path}")
        
        return output_path
    
    def _create_summary_sheet(self, wb: Workbook, data: pd.DataFrame):
        """Create summary statistics sheet"""
        ws = wb.create_sheet("Summary", 0)
        
        # Title
        ws['A1'] = "HVAC Business Intelligence Report"
        ws['A1'].font = Font(size=16, bold=True)
        
        # Generation info
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A4'] = f"Total Businesses Analyzed: {len(data)}"
        
        # Key metrics
        metrics = [
            ("", ""),
            ("KEY METRICS", ""),
            ("Businesses with phone numbers:", len(data[data['phone'] != ""])),
            ("Businesses with owner names:", len(data[data['owner_name'] != ""])),
            ("Businesses with 0 reviews:", len(data[data['review_count'] == 0])),
            ("Businesses with 1-3 reviews:", len(data[data['review_count'].between(1, 3)])),
            ("Businesses with 4-10 reviews:", len(data[data['review_count'].between(4, 10)])),
            ("Average data quality score:", f"{data['data_quality_score'].mean():.1f}%"),
        ]
        
        for i, (label, value) in enumerate(metrics, 6):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            if label and not label.startswith("Businesses"):
                ws[f'A{i}'].font = Font(bold=True)
        
        # Location breakdown
        location_stats = data.groupby('location').agg({
            'name': 'count',
            'review_count': 'mean',
            'data_quality_score': 'mean'
        }).round(1)
        
        ws['A16'] = "LOCATION BREAKDOWN"
        ws['A16'].font = Font(bold=True)
        
        headers = ['Location', 'Count', 'Avg Reviews', 'Avg Quality Score']
        for i, header in enumerate(headers, 1):
            ws.cell(17, i, header).font = Font(bold=True)
        
        for i, (location, stats) in enumerate(location_stats.iterrows(), 18):
            ws.cell(i, 1, location)
            ws.cell(i, 2, stats['name'])
            ws.cell(i, 3, stats['review_count'])
            ws.cell(i, 4, f"{stats['data_quality_score']:.1f}%")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_calling_list_sheet(self, wb: Workbook, data: pd.DataFrame):
        """Create prioritized calling list sheet"""
        ws = wb.create_sheet("Calling List")
        
        # Headers
        headers = [
            'Priority Rank', 'Business Name', 'Location', 'Phone Number', 
            'Owner Name', 'Address', 'Website', 'Review Count', 
            'Star Rating', 'Hours', 'Additional Contact', 'Quality Score'
        ]
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for i, header in enumerate(headers, 1):
            cell = ws.cell(1, i, header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Add data
        for row_idx, (_, row) in enumerate(data.iterrows(), 2):
            ws.cell(row_idx, 1, row_idx - 1)  # Priority rank
            ws.cell(row_idx, 2, row['name'])
            ws.cell(row_idx, 3, row['location'])
            ws.cell(row_idx, 4, row['phone'])
            ws.cell(row_idx, 5, row['owner_name'] or 'Unknown')
            ws.cell(row_idx, 6, row['address'])
            ws.cell(row_idx, 7, row['website'])
            ws.cell(row_idx, 8, row['review_count'])
            ws.cell(row_idx, 9, row['star_rating'])
            ws.cell(row_idx, 10, row['hours'])
            ws.cell(row_idx, 11, row['additional_contact'])
            ws.cell(row_idx, 12, f"{row['data_quality_score']:.1f}%")
            
            # Highlight top 10 priorities
            if row_idx <= 11:  # Top 10 (including header)
                priority_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                for col in range(1, len(headers) + 1):
                    ws.cell(row_idx, col).fill = priority_fill
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_zero_reviews_sheet(self, wb: Workbook, data: pd.DataFrame):
        """Create sheet for businesses with zero reviews"""
        ws = wb.create_sheet("Zero Reviews")
        
        zero_reviews = data[data['review_count'] == 0]
        
        if zero_reviews.empty:
            ws['A1'] = "No businesses found with zero reviews"
            return
        
        # Use same structure as calling list
        headers = [
            'Business Name', 'Location', 'Phone Number', 'Owner Name', 
            'Address', 'Website', 'Hours', 'Additional Contact', 'Quality Score'
        ]
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="DC143C", end_color="DC143C", fill_type="solid")
        
        for i, header in enumerate(headers, 1):
            cell = ws.cell(1, i, header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Add data
        for row_idx, (_, row) in enumerate(zero_reviews.iterrows(), 2):
            ws.cell(row_idx, 1, row['name'])
            ws.cell(row_idx, 2, row['location'])
            ws.cell(row_idx, 3, row['phone'])
            ws.cell(row_idx, 4, row['owner_name'] or 'Unknown')
            ws.cell(row_idx, 5, row['address'])
            ws.cell(row_idx, 6, row['website'])
            ws.cell(row_idx, 7, row['hours'])
            ws.cell(row_idx, 8, row['additional_contact'])
            ws.cell(row_idx, 9, f"{row['data_quality_score']:.1f}%")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_location_breakdown_sheet(self, wb: Workbook, data: pd.DataFrame):
        """Create sheet with breakdown by location"""
        ws = wb.create_sheet("By Location")
        
        current_row = 1
        
        for location in sorted(data['location'].unique()):
            location_data = data[data['location'] == location]
            
            # Location header
            cell = ws.cell(current_row, 1, f"{location} ({len(location_data)} businesses)")
            cell.font = Font(bold=True, size=14)
            current_row += 2
            
            # Headers
            headers = [
                'Business Name', 'Phone Number', 'Owner Name', 'Review Count', 
                'Star Rating', 'Address', 'Website', 'Quality Score'
            ]
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for i, header in enumerate(headers, 1):
                cell = ws.cell(current_row, i, header)
                cell.font = header_font
                cell.fill = header_fill
            current_row += 1
            
            # Data
            for _, row in location_data.iterrows():
                ws.cell(current_row, 1, row['name'])
                ws.cell(current_row, 2, row['phone'])
                ws.cell(current_row, 3, row['owner_name'] or 'Unknown')
                ws.cell(current_row, 4, row['review_count'])
                ws.cell(current_row, 5, row['star_rating'])
                ws.cell(current_row, 6, row['address'])
                ws.cell(current_row, 7, row['website'])
                ws.cell(current_row, 8, f"{row['data_quality_score']:.1f}%")
                current_row += 1
            
            current_row += 2  # Space between locations
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_data_quality_sheet(self, wb: Workbook, data: pd.DataFrame):
        """Create data quality analysis sheet"""
        ws = wb.create_sheet("Data Quality")
        
        # Title
        ws['A1'] = "Data Quality Analysis"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Quality score distribution
        quality_bins = [0, 40, 60, 80, 100]
        quality_labels = ['Poor (0-40%)', 'Fair (40-60%)', 'Good (60-80%)', 'Excellent (80-100%)']
        
        ws['A3'] = "Quality Score Distribution:"
        ws['A3'].font = Font(bold=True)
        
        for i, (bin_start, bin_end, label) in enumerate(zip(quality_bins[:-1], quality_bins[1:], quality_labels), 4):
            count = len(data[(data['data_quality_score'] >= bin_start) & (data['data_quality_score'] < bin_end)])
            ws[f'A{i}'] = label
            ws[f'B{i}'] = count
        
        # Field completeness analysis
        ws['A9'] = "Field Completeness:"
        ws['A9'].font = Font(bold=True)
        
        fields = ['name', 'phone', 'address', 'owner_name', 'website', 'additional_contact']
        field_labels = ['Business Name', 'Phone Number', 'Address', 'Owner Name', 'Website', 'Additional Contact']
        
        for i, (field, label) in enumerate(zip(fields, field_labels), 10):
            if field in data.columns:
                completeness = (data[field] != "").sum() / len(data) * 100
                ws[f'A{i}'] = label
                ws[f'B{i}'] = f"{completeness:.1f}%"
        
        # Top quality businesses
        ws['A17'] = "Highest Quality Businesses:"
        ws['A17'].font = Font(bold=True)
        
        top_quality = data.nlargest(10, 'data_quality_score')
        
        headers = ['Rank', 'Business Name', 'Quality Score', 'Phone', 'Owner']
        for i, header in enumerate(headers, 1):
            ws.cell(18, i, header).font = Font(bold=True)
        
        for i, (_, row) in enumerate(top_quality.iterrows(), 19):
            ws.cell(i, 1, i - 18)
            ws.cell(i, 2, row['name'])
            ws.cell(i, 3, f"{row['data_quality_score']:.1f}%")
            ws.cell(i, 4, row['phone'])
            ws.cell(i, 5, row['owner_name'] or 'Unknown')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def export_csv(self, output_path: str, filtered_data: Optional[pd.DataFrame] = None) -> str:
        """Export data to CSV format"""
        if filtered_data is None:
            filtered_data = self.filter_by_criteria()
        
        # Select and order columns for CSV export
        csv_columns = [
            'name', 'location', 'phone', 'owner_name', 'address', 
            'website', 'review_count', 'star_rating', 'hours', 
            'additional_contact', 'data_quality_score', 'priority_score'
        ]
        
        csv_data = filtered_data[csv_columns].copy()
        csv_data.to_csv(output_path, index=False)
        
        logger.info(f"CSV export saved: {output_path}")
        return output_path
    
    def export_json(self, output_path: str, filtered_data: Optional[pd.DataFrame] = None) -> str:
        """Export data to JSON format"""
        if filtered_data is None:
            filtered_data = self.filter_by_criteria()
        
        # Convert to JSON-friendly format
        json_data = {
            'metadata': {
                'generated_at': datetime.now().isoformat(),
                'total_businesses': len(filtered_data),
                'locations': filtered_data['location'].unique().tolist()
            },
            'businesses': filtered_data.to_dict('records')
        }
        
        with open(output_path, 'w') as f:
            json.dump(json_data, f, indent=2, default=str)
        
        logger.info(f"JSON export saved: {output_path}")
        return output_path

# Example usage
def main():
    """Example usage of the data processor"""
    from hvac_scraper_core import BusinessInfo
    
    # Create sample data
    sample_businesses = [
        BusinessInfo(
            name="Test HVAC Company",
            phone="(208) 555-1234",
            review_count=2,
            star_rating=5.0,
            location="Kuna, Idaho",
            owner_name="John Smith"
        ),
        BusinessInfo(
            name="Another HVAC LLC",
            phone="(208) 555-5678",
            review_count=0,
            star_rating=0.0,
            location="Star, Idaho",
            owner_name=""
        )
    ]
    
    # Process data
    processor = HVACDataProcessor()
    processor.load_businesses(sample_businesses)
    
    # Clean and validate
    cleaned_data = processor.clean_and_validate_data()
    print(f"Cleaned data: {len(cleaned_data)} businesses")
    
    # Filter by criteria
    filtered_data = processor.filter_by_criteria(max_reviews=10)
    print(f"Filtered data: {len(filtered_data)} businesses")
    
    # Generate reports
    processor.generate_excel_report("sample_report.xlsx", filtered_data)
    processor.export_csv("sample_data.csv", filtered_data)
    processor.export_json("sample_data.json", filtered_data)

if __name__ == "__main__":
    main()

